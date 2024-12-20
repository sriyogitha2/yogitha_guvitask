from openpyxl import load_workbook  # Import the load_workbook function from openpyxl to load Excel files

class ExcelFunctions:
    """
    This class provides methods to interact with an Excel file, such as reading, writing, and counting rows/columns.
    """

    def __init__(self, file_name, sheet_name):
        """
        Initializes the ExcelFunctions class with the provided Excel file and sheet name.
        """
        self.file = file_name  # Store the Excel file name
        self.sheet = sheet_name  # Store the sheet name

    def row_count(self):
        """
        Returns the total number of rows in the specified sheet.
        """
        workbook = load_workbook(self.file)  # Load the workbook
        sheet = workbook[self.sheet]  # Access the specified sheet in the workbook
        return sheet.max_row  # Return the maximum number of rows in the sheet
    
    def column_count(self):
        """
        Returns the total number of columns in the specified sheet.
        """
        workbook = load_workbook(self.file)  # Load the workbook
        sheet = workbook[self.sheet]  # Access the specified sheet in the workbook
        return sheet.max_column  # Return the maximum number of columns in the sheet
    
    def read_data(self, row_number, column_number):
        """
        Reads the value from a specified row and column in the sheet.
        """
        workbook = load_workbook(self.file)  # Load the workbook
        sheet = workbook[self.sheet]  # Access the specified sheet in the workbook
        return sheet.cell(row=row_number, column=column_number).value  # Return the value of the cell at (row, column)
    
    def write_data(self, row_number, column_number, data):
        """
        Writes data to a specified cell in the sheet.
        """
        workbook = load_workbook(self.file)  # Load the workbook
        sheet = workbook[self.sheet]  # Access the specified sheet in the workbook
        sheet.cell(row=row_number, column=column_number).value = data  # Write the given data to the specified cell
        workbook.save(self.file)  # Save the workbook after making changes
